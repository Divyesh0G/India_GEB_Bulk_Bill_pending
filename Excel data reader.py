import datetime
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

from tester import infofill


def excel_tableprint(infofill):
    # spl = str(datetime.datetime.now()).split(':')
    shtname = str(datetime.date.today())#spl[0] #+ '-' + spl[1]


    # lode excel file
    wb = openpyxl.load_workbook('hello.xlsx')


    t=wb.get_active_sheet


    # creat a sheet loop

    # if t == shtname:
    #     print(t)
    #     print("its a ,atch")
    # else:
    #     wb.create_sheet(shtname)
    #     wb.active=-1
    #     print("here")

    print(wb.sheetnames)
    print(wb.active)
    ws = wb.active

    ws.append([" Entry No. ", " Flat No. ", " CustomerID ", " Date ", " Amount "])
    for column in infofill:
        ws.append(column)

    tab = Table(displayName="Table1", ref="A1:E5")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save('hello.xlsx')

if __name__ == '__main__':
    excel_tableprint(infofill)